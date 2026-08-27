# -*- coding: utf-8 -*-
"""
网站导航可视化管理工具
功能：导入 Excel / 导入 JSON / 编辑数据 / 搜索筛选 / 分类排序 / 网址排序 / 导出 pintree.json
"""

import json
import os
import sys
import time
import threading
import webbrowser
from collections import OrderedDict
from pathlib import Path
from urllib.parse import urlparse
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog

try:
    import openpyxl
except ImportError:
    print("缺少依赖: openpyxl  请先运行:  pip install openpyxl")
    sys.exit(1)

try:
    import requests
except ImportError:
    requests = None  # 图标下载功能需要 requests，缺失时给出友好提示，不影响其它功能

try:
    import io
    from PIL import Image as PILImage
except ImportError:
    io = None
    PILImage = None  # 图标统一转 PNG 需要 Pillow，缺失时按原始格式保存


# ============== 配置 ==============
HEADERS = ["网站名称", "网址分类目录", "链接", "描述"]
SEPARATOR = ">"          # 分类目录分隔符
SAVE_JSON_NAME = "pintree.json"
APP_TITLE = "网站导航可视化管理工具"
FAVICON_IM = "https://favicon.im/{d}"

# 网站图标本地存放目录：项目根（含 index.html 的目录）下的 assets/logo。
# 兼容多种运行方式：直接运行 .py、在 __pycache__ 中运行 .pyc、PyInstaller 打包后的 exe 等，
# 均从脚本/exe 实际位置逐级向上查找项目根，避免层级固定导致图标落到错误目录。
def _detect_logo_dir():
    if getattr(sys, "frozen", False):          # PyInstaller 打包场景：以 exe 所在目录为基准
        base = os.path.dirname(os.path.abspath(sys.executable))
    else:                                       # 脚本场景：以脚本所在目录为基准
        base = os.path.dirname(os.path.abspath(__file__))
    d = base
    for _ in range(6):                          # 最多向上 6 级查找项目根
        if os.path.isfile(os.path.join(d, "index.html")):
            return os.path.join(d, "assets", "logo")
        nxt = os.path.dirname(d)
        if nxt == d:
            break
        d = nxt
    # 回退：找不到 index.html 时，按「脚本位于 <根>/<子目录>/」的常见结构向上两级
    return os.path.join(os.path.dirname(base), "assets", "logo")


LOGO_DIR = _detect_logo_dir()
try:
    os.makedirs(LOGO_DIR, exist_ok=True)
except Exception:
    pass
ICON_EXTS = (".png", ".jpg", ".gif", ".webp", ".ico")


# ============== 工具函数 ==============
def domain_from_url(url):
    if not url:
        return ""
    try:
        netloc = urlparse(url).netloc
        if not netloc:
            return ""
        if netloc.startswith("www."):
            netloc = netloc[4:]
        return netloc.split(":")[0]
    except Exception:
        return ""


def logo_key(url):
    """从链接提取用于图标文件名的稳定 key：保留 www、去掉端口、转小写。"""
    if not url:
        return ""
    try:
        netloc = urlparse(url).netloc
        if not netloc:
            return ""
        return netloc.split(":")[0].lower()
    except Exception:
        return ""


def _logo_local_rel(url):
    """若本地已存在该站点的图标，返回相对路径（assets/logo/...），否则 None。"""
    key = logo_key(url)
    if not key:
        return None
    for ext in ICON_EXTS:
        p = os.path.join(LOGO_DIR, key + ext)
        if os.path.isfile(p):
            return "assets/logo/" + key + ext
    return None


def _logo_key_exists(key):
    """该 key（logo_key 结果）是否已有本地图标（任意扩展名）。"""
    if not key:
        return False
    return any(os.path.isfile(os.path.join(LOGO_DIR, key + ext)) for ext in ICON_EXTS)


def _cached_icon_usable(icon):
    """校验缓存 icon：指向 assets/logo/ 的本地路径若文件已不存在，视为失效。"""
    if not icon:
        return False
    if icon.startswith("assets/logo/"):
        # 相对路径基于项目根（= LOGO_DIR 向上两级，即 assets 的上一级）
        p = os.path.join(os.path.dirname(os.path.dirname(LOGO_DIR)), icon)
        return os.path.isfile(p)
    return True  # 远程 URL 或其它自定义路径，原样保留


def icon_for(url):
    """图标优先用本地下载的文件，否则回退远程图床，最后回退默认图标。"""
    local = _logo_local_rel(url)
    if local:
        return local
    d = domain_from_url(url)
    return FAVICON_IM.format(d=d) if d else "assets/default-icon.svg"


# ============== 图标下载服务（并入自 website-favicon-downloader） ==============
class FaviconService:
    """批量/单个下载网站 favicon，存到 LOGO_DIR；多公共 API 兜底。"""

    def __init__(self, logo_dir=LOGO_DIR, log_callback=None):
        self.logo_dir = Path(logo_dir)
        self.logo_dir.mkdir(parents=True, exist_ok=True)
        self.session = requests.Session() if requests else None
        if self.session is not None:
            self.session.headers.update({
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                               "(KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
            })
        self.api_endpoints = [
            "https://www.google.com/s2/favicons?domain={}&sz=256",
            "https://api.faviconkit.com/{}/256",
            "https://favicon.yandex.net/favicon/{}/256",
            "https://icons.duckduckgo.com/ip3/{}.ico",
            "https://icon.horse/icon/{}",
        ]
        self.log_callback = log_callback

    def log(self, msg):
        if self.log_callback:
            self.log_callback(msg)
        else:
            print(msg)

    def get_favicon_url(self, domain):
        """尝试多个公共 API 获取图标二进制内容，成功返回 (url, content)。"""
        if self.session is None:
            return None, None
        for api_url in self.api_endpoints:
            try:
                url = api_url.format(domain)
                resp = self.session.get(url, timeout=10)
                if resp.status_code == 200 and len(resp.content) > 0:
                    content = resp.content
                    ctype = resp.headers.get("content-type", "").lower()
                    # 大小不作为过滤条件：小于 70KB 的图标一律正常下载；
                    # 仅当响应既不是图片类型、内容也不像图片二进制时才跳过（防错误页）
                    if "image" in ctype or "octet-stream" in ctype or self._is_image_bytes(content):
                        return url, content
                    self.log(f"跳过非图片内容: {domain} ({ctype or '未知类型'}, {len(content)} 字节)")
            except Exception as e:
                self.log(f"API 失败 {api_url.format(domain)}: {e}")
                continue
        return None, None

    def _is_image_bytes(self, content):
        """按文件头魔数判断内容是否为常见图片格式（含 SVG）。"""
        if not content:
            return False
        if content[:4] == b"\x89PNG":
            return True
        if content[:3] == b"\xff\xd8\xff":
            return True
        if content[:6] in (b"GIF87a", b"GIF89a"):
            return True
        if content[:4] in (b"\x00\x00\x01\x00", b"\x00\x00\x02\x00"):
            return True
        if content[:4] == b"RIFF" and content[8:12] == b"WEBP":
            return True
        if content[:2] == b"BM":
            return True
        head = content[:256].lstrip().lower()
        if head.startswith(b"<?xml") or head.startswith(b"<svg") or b"<svg" in head:
            return True
        return False

    def ext_for_data(self, content):
        """按文件头判断图像扩展名（增强：含 .ico 识别）。"""
        if content[:4] == b"\x89PNG":
            return ".png"
        if content[:3] == b"\xff\xd8\xff":
            return ".jpg"
        if content[:6] in (b"GIF87a", b"GIF89a"):
            return ".gif"
        if content[:4] in (b"\x00\x00\x01\x00", b"\x00\x00\x02\x00"):
            return ".ico"
        if content[:4] == b"RIFF" and content[8:12] == b"WEBP":
            return ".webp"
        return ".png"

    def save_as_png(self, key, content):
        """将图标内容统一转换为 PNG 保存为 {key}.png，返回最终文件名。
        Pillow 缺失或无法识别（如 SVG）时按原始格式回退保存。"""
        # 已是 PNG 直接落盘
        if content[:4] == b"\x89PNG":
            path = self.logo_dir / f"{key}.png"
            with open(path, "wb") as f:
                f.write(content)
            return path.name
        # 尝试用 Pillow 转成 PNG
        if PILImage is not None:
            try:
                img = PILImage.open(io.BytesIO(content))
                if img.mode not in ("RGBA", "RGB"):
                    img = img.convert("RGBA")
                path = self.logo_dir / f"{key}.png"
                img.save(path, "PNG")
                return path.name
            except Exception:
                pass
        # 回退：按原始格式保存（SVG 等无法转换的情况）
        ext = self.ext_for_data(content)
        path = self.logo_dir / f"{key}{ext}"
        with open(path, "wb") as f:
            f.write(content)
        return path.name

    def download_one(self, key):
        """下载单个站点图标（key = logo_key(url)），统一转为 PNG，成功返回 (True, url)。
        转换后的 PNG 若 ≤ 70 字节（无效/空图标），不保留，视为下载失败。"""
        favicon_url, content = self.get_favicon_url(key)
        if not content:
            self.log(f"无图标: {key}")
            return False, None
        try:
            self.remove_existing(key)  # 先清除旧文件，避免多扩展名并存
            name = self.save_as_png(key, content)
            path = self.logo_dir / name
            try:
                size = path.stat().st_size if path.is_file() else len(content)
            except OSError:
                size = len(content)
            if size <= 70:
                try:
                    if path.is_file():
                        path.unlink()
                except OSError:
                    pass
                self.log(f"丢弃: {key} 转换后图标过小（{size} 字节 ≤ 70），未保存")
                return False, favicon_url
            self.log(f"成功: {key} -> {name}")
            return True, favicon_url
        except Exception as e:
            self.log(f"保存失败 {key}: {e}")
            return False, favicon_url

    def remove_existing(self, key):
        """删除该 key 已有的任意扩展名图标，避免多个文件并存。"""
        for ext in ICON_EXTS:
            p = self.logo_dir / f"{key}{ext}"
            try:
                if p.is_file():
                    p.unlink()
            except Exception:
                pass


# ============== 数据模型 ==============
class DataStore:
    def __init__(self):
        self.items = []              # [dict{name, category, url, desc}]   # 顺序 = 显示顺序
        self.current_file = None     # 已打开或保存的 xlsx 文件路径
        self.modified = False

    # --- Excel 读写 ---
    def load_excel(self, path):
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        items = []
        for r in range(2, ws.max_row + 1):
            name = ws.cell(row=r, column=1).value or ""
            cat = ws.cell(row=r, column=2).value or ""
            url = ws.cell(row=r, column=3).value or ""
            desc = ws.cell(row=r, column=4).value or ""
            if not name and not url and not cat:
                continue
            items.append({
                "name": str(name).strip(),
                "category": str(cat).strip(),
                "url": str(url).strip(),
                "desc": str(desc).strip(),
            })
        self.items = items
        self.current_file = path
        self.modified = False
        return len(items)

    # --- JSON 读取 ---
    def load_json(self, path):
        """读取 pintree.json 嵌套树结构，展开为扁平 items 列表"""
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        items = []

        def walk(nodes, path_parts):
            for node in nodes:
                if node.get("type") == "folder":
                    title = node.get("title", "")
                    new_path = path_parts + [title] if title else path_parts
                    walk(node.get("children", []), new_path)
                elif node.get("type") == "link":
                    category = SEPARATOR.join(path_parts) if path_parts else ""
                    items.append({
                        "name": node.get("title", ""),
                        "category": category,
                        "url": node.get("url", ""),
                        "desc": node.get("description", ""),
                        "icon": node.get("icon", ""),  # 保留原图标地址，下载成功后会改写为本地路径
                    })

        walk(data, [])
        self.items = items
        self.current_file = None  # JSON 不绑定 xlsx 路径，保存时需另存
        self.modified = False
        return len(items)

    def save_excel(self, path=None):
        path = path or self.current_file
        if not path:
            raise ValueError("未指定保存路径")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "网站导航"
        for col_idx, h in enumerate(HEADERS, start=1):
            ws.cell(row=1, column=col_idx, value=h)
        for i, it in enumerate(self.items, start=2):
            ws.cell(row=i, column=1, value=it["name"])
            ws.cell(row=i, column=2, value=it["category"])
            ws.cell(row=i, column=3, value=it["url"])
            ws.cell(row=i, column=4, value=it["desc"])
        wb.save(path)
        self.current_file = path
        self.modified = False
        return path

    # --- 数据操作 ---
    def add(self, item, index=None):
        if index is None or index >= len(self.items):
            self.items.append(item)
        else:
            self.items.insert(index, item)
        self.modified = True

    def update(self, index, item):
        self.items[index] = item
        self.modified = True

    def delete(self, indices):
        for i in sorted(indices, reverse=True):
            if 0 <= i < len(self.items):
                self.items.pop(i)
        self.modified = True

    def move_up(self, indices):
        indices = sorted(set(indices))
        for i in indices:
            if i > 0:
                self.items[i - 1], self.items[i] = self.items[i], self.items[i - 1]
        self.modified = True

    def move_down(self, indices):
        indices = sorted(set(indices), reverse=True)
        for i in indices:
            if i < len(self.items) - 1:
                self.items[i + 1], self.items[i] = self.items[i], self.items[i + 1]
        self.modified = True

    def move_top(self, indices):
        indices = sorted(set(indices))
        new_list = [self.items[i] for i in indices] + [self.items[i] for i in range(len(self.items)) if i not in indices]
        self.items = new_list
        self.modified = True

    def move_bottom(self, indices):
        indices = sorted(set(indices))
        new_list = [self.items[i] for i in range(len(self.items)) if i not in indices] + [self.items[i] for i in indices]
        self.items = new_list
        self.modified = True

    # --- 分类统计 ---
    def category_counts(self):
        """返回 OrderedDict { 分类完整路径字符串 -> 数量 }，按首次出现顺序"""
        counts = OrderedDict()
        for it in self.items:
            cat = it["category"] or "未分类"
            counts[cat] = counts.get(cat, 0) + 1
        # 聚合父分类计数: "a>b" 的每个前缀都要计入
        agg = OrderedDict()

        def incr(key):
            agg[key] = agg.get(key, 0) + 1

        for it in self.items:
            cat = it["category"] or "未分类"
            parts = [p.strip() for p in cat.split(SEPARATOR)]
            key = ""
            for j, p in enumerate(parts):
                key = p if j == 0 else key + SEPARATOR + p
                incr(key)
        return counts, agg

    # --- 分类树结构（用于导出 JSON） ---
    def to_nested_tree(self):
        """按 items 顺序构建嵌套 folder/link 树结构，返回顶级 folder 列表"""
        base_ts = 1718526477999
        counter = [0]

        def nxt():
            counter[0] += 1
            return base_ts + counter[0]

        def make_folder(t, ch):
            return {"type": "folder", "addDate": nxt(), "title": t, "children": ch}

        def make_link(it):
            # 优先本地图标（磁盘存在）；其次导入时缓存的 icon（本地路径需校验文件仍存在，
            # 已删除的视为失效并丢弃，回退图床）；最后回退图床
            cached = it.get("icon") or ""
            if not _cached_icon_usable(cached):
                cached = ""
            icon = _logo_local_rel(it.get("url", "")) or cached or icon_for(it["url"])
            return {
                "type": "link",
                "addDate": nxt(),
                "title": it["name"] or it["url"],
                "icon": icon,
                "url": it["url"],
                "description": it["desc"],
            }

        # 顶层字典：一级分类名 -> (folder_node, top_order_index)
        top_order = []
        top_map = {}

        for it in self.items:
            cat = it["category"] or "未分类"
            parts = [p.strip() for p in cat.split(SEPARATOR) if p.strip()]
            if not parts:
                parts = ["未分类"]

            top_name = parts[0]
            if top_name not in top_map:
                f = make_folder(top_name, [])
                top_map[top_name] = f
                top_order.append(f)
            folder_node = top_map[top_name]

            # 下钻中间级别
            for i in range(1, len(parts)):
                sub_name = parts[i]
                sub_node = None
                for ch in folder_node["children"]:
                    if ch.get("type") == "folder" and ch.get("title") == sub_name:
                        sub_node = ch
                        break
                if sub_node is None:
                    sub_node = make_folder(sub_name, [])
                    folder_node["children"].append(sub_node)
                folder_node = sub_node

            folder_node["children"].append(make_link(it))

        return top_order


# ============== 主窗口 ==============
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1280x780")
        self.minsize(1100, 650)
        self.store = DataStore()
        self._build_style()
        self._build_ui()
        self._refresh_all()

    # -------- 样式 --------
    def _build_style(self):
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except Exception:
            pass
        style.configure("TButton", padding=6)
        style.configure("Toolbar.TButton", padding=(10, 6))
        style.configure("Sort.TButton", padding=(8, 4))
        style.configure("TLabelframe", padding=6)
        style.configure("Treeview", rowheight=24)
        style.configure("Treeview.Heading", font=("Microsoft YaHei", 9, "bold"))
        style.configure("TLabelframe.Label", font=("Microsoft YaHei", 9))

    # -------- UI --------
    def _build_ui(self):
        def btn(parent, text, cmd, style="TButton", width=None):
            kw = {"text": text, "command": cmd, "style": style}
            if width is not None:
                kw["width"] = width
            return ttk.Button(parent, **kw)

        # ========== 第一行：文件/编辑主工具栏 ==========
        toolbar = ttk.Frame(self, padding=6)
        toolbar.pack(side=tk.TOP, fill=tk.X)

        file_box = ttk.LabelFrame(toolbar, text="文件", padding=4)
        file_box.pack(side=tk.LEFT, padx=(0, 6))
        btn(file_box, "导入 Excel", self.on_import_excel, style="Toolbar.TButton").pack(side=tk.LEFT, padx=2)
        btn(file_box, "导入 JSON", self.on_import_json, style="Toolbar.TButton").pack(side=tk.LEFT, padx=2)
        btn(file_box, "保存", self.on_save, style="Toolbar.TButton").pack(side=tk.LEFT, padx=2)
        btn(file_box, "另存为...", self.on_save_as, style="Toolbar.TButton").pack(side=tk.LEFT, padx=2)
        self.btn_export = btn(file_box, "导出 json", self.on_export_json, style="Toolbar.TButton")
        self.btn_export.pack(side=tk.LEFT, padx=2)

        edit_box = ttk.LabelFrame(toolbar, text="编辑", padding=4)
        edit_box.pack(side=tk.LEFT, padx=6)
        btn(edit_box, "新增", self.on_add, style="Toolbar.TButton").pack(side=tk.LEFT, padx=2)
        btn(edit_box, "编辑", self.on_edit, style="Toolbar.TButton").pack(side=tk.LEFT, padx=2)
        btn(edit_box, "删除", self.on_delete, style="Toolbar.TButton").pack(side=tk.LEFT, padx=2)
        btn(edit_box, "打开链接", self.on_open_link, style="Toolbar.TButton").pack(side=tk.LEFT, padx=2)

        # ========== 图标操作 ==========
        icon_box = ttk.LabelFrame(toolbar, text="图标", padding=4)
        icon_box.pack(side=tk.LEFT, padx=6)
        btn(icon_box, "下载图标", self.on_download_icons, style="Toolbar.TButton").pack(side=tk.LEFT, padx=2)
        btn(icon_box, "更新图标", self.on_update_icon, style="Toolbar.TButton").pack(side=tk.LEFT, padx=2)

        # ========== 第二行：排序/搜索工具栏 ==========
        sort_bar = ttk.Frame(self, padding=6)
        sort_bar.pack(side=tk.TOP, fill=tk.X)

        # 网址排序
        link_sort_box = ttk.LabelFrame(sort_bar, text="网址排序", padding=6)
        link_sort_box.pack(side=tk.LEFT, padx=(0, 8))
        self.sort_scope = tk.StringVar(value="cur")        # cur=当前筛选 / same=同分类 / all=全部
        ttk.Radiobutton(link_sort_box, text="当前筛选", variable=self.sort_scope, value="cur").pack(side=tk.LEFT)
        ttk.Radiobutton(link_sort_box, text="同分类", variable=self.sort_scope, value="same").pack(side=tk.LEFT, padx=6)
        ttk.Radiobutton(link_sort_box, text="全部", variable=self.sort_scope, value="all").pack(side=tk.LEFT)
        ttk.Separator(link_sort_box, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=6)
        btn(link_sort_box, "置顶", self.on_link_top, style="Sort.TButton", width=6).pack(side=tk.LEFT, padx=2)
        btn(link_sort_box, "上移", self.on_link_up, style="Sort.TButton", width=6).pack(side=tk.LEFT, padx=2)
        btn(link_sort_box, "下移", self.on_link_down, style="Sort.TButton", width=6).pack(side=tk.LEFT, padx=2)
        btn(link_sort_box, "置底", self.on_link_bottom, style="Sort.TButton", width=6).pack(side=tk.LEFT, padx=2)

        # 分类排序
        cat_sort_box = ttk.LabelFrame(sort_bar, text="分类排序", padding=6)
        cat_sort_box.pack(side=tk.LEFT, padx=(0, 8))
        btn(cat_sort_box, "置顶", self.on_cat_top, style="Sort.TButton", width=6).grid(row=0, column=0, padx=2)
        btn(cat_sort_box, "上移", self.on_cat_up, style="Sort.TButton", width=6).grid(row=0, column=1, padx=2)
        btn(cat_sort_box, "下移", self.on_cat_down, style="Sort.TButton", width=6).grid(row=0, column=2, padx=2)
        btn(cat_sort_box, "置底", self.on_cat_bottom, style="Sort.TButton", width=6).grid(row=0, column=3, padx=2)
        more_btn = ttk.Menubutton(cat_sort_box, text="更多 ▾", direction="below", width=8)
        more_btn.grid(row=0, column=4, padx=(6, 2))
        more_menu = tk.Menu(more_btn, tearoff=False)
        more_btn.config(menu=more_menu)
        more_menu.add_command(label="按名称 A→Z（正序）", command=lambda: self.on_cat_sort("name", 1))
        more_menu.add_command(label="按名称 Z→A（倒序）", command=lambda: self.on_cat_sort("name", -1))
        more_menu.add_command(label="按数量 少→多（正序）", command=lambda: self.on_cat_sort("count", 1))
        more_menu.add_command(label="按数量 多→少（倒序）", command=lambda: self.on_cat_sort("count", -1))
        more_menu.add_separator()
        more_menu.add_command(label="分类内网址 按名称排序 正序", command=lambda: self.on_links_in_cat_sort("name", 1))
        more_menu.add_command(label="分类内网址 按名称排序 倒序", command=lambda: self.on_links_in_cat_sort("name", -1))
        more_menu.add_command(label="分类内网址 按链接排序 正序", command=lambda: self.on_links_in_cat_sort("url", 1))
        more_menu.add_command(label="分类内网址 按链接排序 倒序", command=lambda: self.on_links_in_cat_sort("url", -1))
        more_menu.add_separator()
        more_menu.add_command(label="按表格顺序恢复（原始顺序）", command=self.on_cat_original)

        # 搜索框
        search_frame = ttk.LabelFrame(sort_bar, text="搜索", padding=6)
        search_frame.pack(side=tk.RIGHT, padx=(8, 0), fill=tk.Y)
        ttk.Label(search_frame, text="关键词:").pack(side=tk.LEFT)
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", lambda *a: self._refresh_table())
        self.search_entry = ttk.Entry(search_frame, textvariable=self.search_var, width=26)
        self.search_entry.pack(side=tk.LEFT, padx=6)
        btn(search_frame, "重置", self.on_reset_filter, style="Sort.TButton", width=6).pack(side=tk.LEFT)

        # ========== 主内容：PanedWindow，左=分类树，右=表格 ==========
        # 使用 tk.PanedWindow 而非 ttk.PanedWindow，因为前者支持 minsize 且 sash 更稳定
        self.body = tk.PanedWindow(self, orient=tk.HORIZONTAL, sashrelief=tk.RAISED, sashwidth=4)
        self.body.pack(fill=tk.BOTH, expand=True, padx=6, pady=(2, 0))

        # --- 左侧：分类目录 ---
        left_frame = tk.Frame(self.body, width=280, bg="#f5f5f5")
        left_head = ttk.Frame(left_frame)
        left_head.pack(fill=tk.X, pady=(4, 2))
        ttk.Label(left_head, text="分类目录（含数量）", font=("Microsoft YaHei", 9, "bold")).pack(side=tk.LEFT)
        ttk.Label(left_head, text="（右侧按钮可调整顺序）", foreground="#666").pack(side=tk.LEFT, padx=6)

        self.cat_tree = ttk.Treeview(left_frame, show="tree")
        cat_scroll = ttk.Scrollbar(left_frame, orient=tk.VERTICAL, command=self.cat_tree.yview)
        self.cat_tree.configure(yscrollcommand=cat_scroll.set)
        self.cat_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, pady=2)
        cat_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.cat_tree.bind("<<TreeviewSelect>>", self.on_cat_select)

        self.body.add(left_frame, minsize=240)

        # --- 右侧：数据表格 ---
        right_frame2 = tk.Frame(self.body, bg="#f5f5f5")
        # 统计条
        self.count_label = ttk.Label(right_frame2, text="显示 0 / 0 条", foreground="#555")
        self.count_label.pack(anchor=tk.W, pady=(4, 2))
        # 表格
        table_frame = ttk.Frame(right_frame2)
        table_frame.pack(fill=tk.BOTH, expand=True)
        self.table = ttk.Treeview(table_frame, columns=("name", "category", "url", "desc"),
                                  show="headings", selectmode="extended")
        for col, h, w in [("name", "网站名称", 220), ("category", "网址分类目录", 180),
                          ("url", "链接", 320), ("desc", "描述", 520)]:
            self.table.heading(col, text=h)
            self.table.column(col, width=w, minwidth=100, anchor=tk.W, stretch=True)
        vscr = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.table.yview)
        hscr = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=self.table.xview)
        self.table.configure(yscrollcommand=vscr.set, xscrollcommand=hscr.set)
        self.table.grid(row=0, column=0, sticky="nsew")
        vscr.grid(row=0, column=1, sticky="ns")
        hscr.grid(row=1, column=0, sticky="ew")
        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)
        self.table.bind("<Double-1>", lambda e: self.on_edit())
        self.body.add(right_frame2, minsize=500)

        # --- 状态栏 ---
        self.status_var = tk.StringVar(value="就绪。点击'导入 Excel'开始，或'导出 json'生成 pintree.json。")
        status = ttk.Label(self, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W, padding=(8, 3))
        status.pack(side=tk.BOTTOM, fill=tk.X)

        self._data_rowids = []  # 当前显示项（筛选后）对应 store.items 的下标列表

    # ============== 刷新 ==============
    def _refresh_all(self):
        self._refresh_categories()
        self._refresh_table()

    def _cat_root_items(self):
        """返回顶级分类在 store 中首次出现的顺序"""
        seen = OrderedDict()
        for i, it in enumerate(self.store.items):
            cat = it["category"] or "未分类"
            parts = [p.strip() for p in cat.split(SEPARATOR) if p.strip()]
            root = parts[0] if parts else "未分类"
            if root not in seen:
                seen[root] = True
        return list(seen.keys())

    def _save_layout(self):
        """保存窗口几何与分栏位置，供刷新后恢复，防止界面收缩"""
        geo = self.geometry()
        sash_pos = None
        try:
            sash_pos = self.body.sash_coord(0)[0]
        except Exception:
            pass
        return geo, sash_pos

    def _restore_layout(self, geo=None, sash_pos=None):
        """刷新后恢复窗口几何与分栏位置。
        用 after_idle 延迟执行，等待本次刷新触发的布局全部完成后再恢复，
        否则 sash 位置会被后续布局阶段覆盖（表现为左侧面板收缩）。
        """
        def _apply():
            if geo:
                try:
                    self.geometry(geo)
                except Exception:
                    pass
            if sash_pos is not None:
                try:
                    self.update_idletasks()
                    self.body.sash_place(0, sash_pos, 0)
                except Exception:
                    pass
        self.after_idle(_apply)
        # 再追加一次延迟恢复，双保险：有些场景下两轮 idle 才稳定
        self.after(120, _apply)

    def _refresh_categories(self):
        # 保存当前选中
        sel = self.cat_tree.selection()
        sel_val = self.cat_tree.item(sel[0], "values")[0] if sel else None

        # 保存展开状态（用分类完整路径作为稳定键）
        open_set = set()
        def collect_open(node_iid):
            for child in self.cat_tree.get_children(node_iid):
                vals = self.cat_tree.item(child, "values")
                if vals and self.cat_tree.item(child, "open"):
                    open_set.add(vals[0])
                collect_open(child)
        collect_open("")

        # 保存 sash 位置，防止刷新后左侧面板收缩
        sash_pos = self._save_layout()[1]

        self.cat_tree.delete(*self.cat_tree.get_children())

        # 全部站点
        total = len(self.store.items)
        root_iid = self.cat_tree.insert(
            "", tk.END, text=f"全部站点 ({total})",
            values=("__ALL__",), open=True, tags=("cat",)
        )

        # 聚合计数
        _, agg = self.store.category_counts()

        # 顶级分类顺序
        top_roots = self._cat_root_items()

        def insert_node(parent_iid, parent_prefix, level_names_left, order_list):
            for name in order_list:
                full = name if not parent_prefix else parent_prefix + SEPARATOR + name
                cnt = agg.get(full, 0)
                # 判断该 full 是否拥有子项(直接属于它+孙子)
                # 存在后代前缀 full+SEPARATOR 就说明有子文件夹
                has_children = any(k.startswith(full + SEPARATOR) for k in agg.keys())
                node_iid = self.cat_tree.insert(
                    parent_iid, tk.END,
                    text=f"{'  ' if parent_iid != '' else ''}  {name} ({cnt})",
                    values=(full,)
                )
                if has_children:
                    # 取属于该 full 的直接下一级名字
                    children_names = OrderedDict()
                    for k in agg.keys():
                        if k.startswith(full + SEPARATOR):
                            rest = k[len(full + SEPARATOR):]
                            first = rest.split(SEPARATOR)[0]
                            children_names[first] = True
                    insert_node(node_iid, full, [], list(children_names.keys()))

        insert_node(root_iid, "", [], top_roots)

        # 展开 root
        if root_iid:
            self.cat_tree.item(root_iid, open=True)

        # 恢复展开状态
        def apply_open(node_iid):
            for child in self.cat_tree.get_children(node_iid):
                vals = self.cat_tree.item(child, "values")
                if vals and vals[0] in open_set:
                    self.cat_tree.item(child, open=True)
                apply_open(child)
        apply_open("")

        # 恢复选中（递归搜索任意层级，并展开祖先节点保证选中项可见）
        if sel_val:
            def find_and_select(node_iid):
                for child in self.cat_tree.get_children(node_iid):
                    try:
                        vals = self.cat_tree.item(child, "values")
                    except Exception:
                        continue
                    if vals and vals[0] == sel_val:
                        return [child]
                    found = find_and_select(child)
                    if found:
                        # 展开当前祖先节点，让深层选中项保持可见
                        try:
                            self.cat_tree.item(child, open=True)
                        except Exception:
                            pass
                        return [child] + found
                return None
            path = find_and_select("")
            if path:
                target = path[-1]
                self.cat_tree.selection_set(target)
                self.cat_tree.see(target)

        # 恢复 sash 位置，保持左侧面板宽度
        if sash_pos is not None:
            self._restore_layout(sash_pos=sash_pos)

    def _refresh_table(self):
        # 筛选：分类 + 搜索关键词
        cat_sel = self.cat_tree.selection()
        cat_filter = None
        if cat_sel:
            v = self.cat_tree.item(cat_sel[0], "values")
            if v and v[0] != "__ALL__":
                cat_filter = v[0]

        kw = self.search_var.get().strip().lower()

        self.table.delete(*self.table.get_children())
        self._data_rowids = []

        for idx, it in enumerate(self.store.items):
            cat = it["category"] or "未分类"
            if cat_filter:
                # cat= "a>b"，筛选 a 时应命中；筛选 a>b 时应命中；筛选 a>b>c 时仅 a>b>c 命中
                if not (cat == cat_filter or cat.startswith(cat_filter + SEPARATOR)):
                    continue
            if kw:
                if (kw not in it["name"].lower()
                        and kw not in it["url"].lower()
                        and kw not in it["desc"].lower()
                        and kw not in it["category"].lower()):
                    continue
            rowid = self.table.insert(
                "", tk.END,
                values=(it["name"], it["category"], it["url"], it["desc"])
            )
            self._data_rowids.append((idx, rowid))

        total = len(self.store.items)
        shown = len(self._data_rowids)
        self.count_label.config(text=f"显示 {shown} / {total} 条")

    # ============== 工具栏动作 ==============
    def on_import_excel(self):
        path = filedialog.askopenfilename(
            title="选择 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx *.xlsm"), ("所有文件", "*.*")],
        )
        if not path:
            return
        try:
            cnt = self.store.load_excel(path)
        except Exception as e:
            messagebox.showerror("导入失败", str(e))
            return
        self._refresh_all()
        self.set_status(f"已导入 {path} （共 {cnt} 条）")
        # 默认选中"全部站点"
        root = self.cat_tree.get_children()
        if root:
            self.cat_tree.selection_set(root[0])

    def on_import_json(self):
        path = filedialog.askopenfilename(
            title="选择 JSON 文件",
            filetypes=[("JSON 文件", "*.json"), ("所有文件", "*.*")],
        )
        if not path:
            return
        try:
            cnt = self.store.load_json(path)
        except Exception as e:
            messagebox.showerror("导入失败", str(e))
            return
        self._refresh_all()
        self.set_status(f"已导入 {path} （共 {cnt} 条）")
        # 默认选中"全部站点"
        root = self.cat_tree.get_children()
        if root:
            self.cat_tree.selection_set(root[0])

    def on_save(self):
        if not self.store.current_file:
            self.on_save_as()
            return
        try:
            p = self.store.save_excel()
            self.set_status(f"已保存到 {p}")
        except Exception as e:
            messagebox.showerror("保存失败", str(e))

    def on_save_as(self):
        init = self.store.current_file or "网站导航.xlsx"
        init_dir = os.path.dirname(init) if os.path.dirname(init) else os.getcwd()
        init_name = os.path.basename(init) or "网站导航.xlsx"
        path = filedialog.asksaveasfilename(
            title="另存为 Excel",
            defaultextension=".xlsx",
            initialdir=init_dir,
            initialfile=init_name,
            filetypes=[("Excel 文件", "*.xlsx")],
        )
        if not path:
            return
        try:
            p = self.store.save_excel(path)
            self.set_status(f"已另存为 {p}")
        except Exception as e:
            messagebox.showerror("保存失败", str(e))

    # ---- 导出 JSON ----
    def on_export_json(self):
        if not self.store.items:
            messagebox.showwarning("没有数据", "请先导入 Excel 或添加数据。")
            return
        # 默认输出路径:与 xlsx 同目录下的 json/pintree.json；若无则当前程序目录下 json/pintree.json
        default_dir = ""
        if self.store.current_file:
            default_dir = os.path.dirname(self.store.current_file)
        if not default_dir:
            default_dir = os.getcwd()
        default_path = os.path.join(default_dir, "json", SAVE_JSON_NAME)

        path = filedialog.asksaveasfilename(
            title="导出 pintree.json",
            defaultextension=".json",
            initialdir=os.path.dirname(default_path),
            initialfile=SAVE_JSON_NAME,
            filetypes=[("JSON 文件", "*.json")],
        )
        if not path:
            return
        try:
            os.makedirs(os.path.dirname(path), exist_ok=True)
            data = self.store.to_nested_tree()
            with open(path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            messagebox.showerror("导出失败", str(e))
            return

        # 统计（含本地图标引用情况）
        total_links = 0
        total_folders = 0
        local_icons = 0

        def walk(nodes):
            nonlocal total_links, total_folders, local_icons
            for n in nodes:
                if n.get("type") == "link":
                    total_links += 1
                    if str(n.get("icon", "")).startswith("assets/logo/"):
                        local_icons += 1
                elif n.get("type") == "folder":
                    total_folders += 1
                    walk(n.get("children", []))

        walk(data)
        self.set_status(
            f"已导出 {path} （{len(data)} 个顶级文件夹 / {total_folders} 总文件夹 / {total_links} 条链接 / "
            f"{local_icons} 条已更新为本地图标）"
        )
        messagebox.showinfo(
            "导出成功",
            f"已成功生成:\n{path}\n\n顶级文件夹: {len(data)}  总文件夹: {total_folders}  链接: {total_links}\n"
            f"图标地址已更新：{local_icons} 条引用本地 PNG（assets/logo/），"
            f"其余仍为远程地址。",
        )

    # ---- 增删改 ----
    def _sel_store_indices(self, scope=None):
        """根据 scope(cur/same/all) 返回 store 中选中项的 index 列表"""
        scope = scope or self.sort_scope.get()
        sel_rows = self.table.selection()
        rowid_to_idx = {r: i for i, r in self._data_rowids}
        if scope == "cur":
            # 当前显示项中的选中
            return [rowid_to_idx[r] for r in sel_rows if r in rowid_to_idx]
        if scope == "all":
            # 全部
            return [rowid_to_idx[r] for r in sel_rows if r in rowid_to_idx]
        if scope == "same":
            # 同分类：取选中项的分类，返回该分类在 store 中所有选中的条目
            idxs = [rowid_to_idx[r] for r in sel_rows if r in rowid_to_idx]
            if not idxs:
                return []
            cat = self.store.items[idxs[0]]["category"]
            return [i for i in idxs if self.store.items[i]["category"] == cat]
        return []

    def on_add(self):
        dlg = ItemDialog(self, title="新增网站")
        if dlg.result:
            self.store.add(dlg.result)
            self._refresh_all()
            # 新分类可能改变树结构，保持选中全部站点便于查看
            root = self.cat_tree.get_children()
            if root:
                self.cat_tree.selection_set(root[0])
            self.set_status("已新增 1 条。")

    def on_edit(self):
        idxs = self._sel_store_indices()
        if len(idxs) != 1:
            messagebox.showinfo("提示", "请在右侧表格中选择一行进行编辑。")
            return
        dlg = ItemDialog(self, title="编辑网站", item=self.store.items[idxs[0]])
        if dlg.result:
            self.store.update(idxs[0], dlg.result)
            self._refresh_all()
            self.set_status("已编辑。")

    def on_delete(self):
        idxs = self._sel_store_indices()
        if not idxs:
            messagebox.showinfo("提示", "请先选择要删除的行。")
            return
        if not messagebox.askyesno("确认", f"确定删除选中的 {len(idxs)} 条记录？"):
            return
        self.store.delete(idxs)
        self._refresh_all()
        self.set_status(f"已删除 {len(idxs)} 条。")

    def on_open_link(self):
        idxs = self._sel_store_indices()
        if not idxs:
            messagebox.showinfo("提示", "请先选择要打开的行。")
            return
        ok = 0
        for i in idxs:
            url = self.store.items[i]["url"]
            if url and (url.startswith("http://") or url.startswith("https://")):
                webbrowser.open(url)
                ok += 1
        self.set_status(f"已尝试在浏览器中打开 {ok} 个链接。")

    # ---- 网址排序 ----
    def _apply_link_sort(self, func):
        idxs = self._sel_store_indices()
        if not idxs:
            messagebox.showinfo("提示", "请先在右侧表格选中要排序的条目。")
            return
        # 保存窗口几何与 sash 位置
        geo, sash_pos = self._save_layout()
        func(idxs)
        self._refresh_all()
        # 恢复窗口几何与 sash 位置
        self._restore_layout(geo, sash_pos)
        self.set_status("已调整顺序。")

    def on_link_top(self):
        self._apply_link_sort(self.store.move_top)

    def on_link_bottom(self):
        self._apply_link_sort(self.store.move_bottom)

    def on_link_up(self):
        self._apply_link_sort(self.store.move_up)

    def on_link_down(self):
        self._apply_link_sort(self.store.move_down)

    # ---- 分类排序（支持任意级分类（顶级 / 子级）顺序调整） ----
    def _selected_cat_fullpath(self):
        sel = self.cat_tree.selection()
        if not sel:
            return None
        v = self.cat_tree.item(sel[0], "values")
        if not v or v[0] == "__ALL__":
            return None
        return v[0]

    def _sibling_categories(self, target_full):
        """
        返回 target_full 所属的同级分类列表（完整路径字符串），以及它自身在列表中的下标。
        如 target='实用工具>图片处理' → 返回属于父'实用工具'下所有直接子分类（完整路径）的有序列表，以及 idx。
        顶级时，parent_prefix=''，返回顶级分类完整路径(=分类名)列表。
        """
        if SEPARATOR in target_full:
            parent_prefix = target_full.rsplit(SEPARATOR, 1)[0]
            target_short = target_full[len(parent_prefix + SEPARATOR):]
        else:
            parent_prefix = ""
            target_short = target_full

        # 收集所有分类出现顺序中，同级兄弟（去重）
        siblings = OrderedDict()
        for it in self.store.items:
            cat = it["category"] or "未分类"
            if parent_prefix:
                if not cat.startswith(parent_prefix + SEPARATOR):
                    continue
                rest = cat[len(parent_prefix + SEPARATOR):]
                if not rest:
                    continue
                short = rest.split(SEPARATOR)[0].strip()
            else:
                root = cat.split(SEPARATOR)[0].strip()
                short = root
            siblings[short] = True

        sibling_fulls = [
            (short if not parent_prefix else parent_prefix + SEPARATOR + short)
            for short in siblings.keys()
        ]
        try:
            idx = sibling_fulls.index(target_full)
        except ValueError:
            idx = -1
        return sibling_fulls, idx, parent_prefix, target_short

    def _reorder_sibling_categories(self, sibling_fulls_new, parent_prefix):
        """
        按 sibling_fulls_new 中同级分类顺序重排 store.items 中兄弟项。
        parent_prefix='' 时处理顶级；='a>b' 时处理 'a>b>xxx'。
        组内其它条目的相对顺序保持不变；非同级其它分类的原有整体顺序保持。
        """
        # 保存窗口几何，防止刷新触发收缩
        geo, sash_pos = self._save_layout()

        if parent_prefix:
            prefix_filter = parent_prefix + SEPARATOR
            sibling_shorts_new = [f[len(prefix_filter):] for f in sibling_fulls_new]
            # key = 同级短名，value=所属 bucket
            buckets = OrderedDict()
            for s in sibling_shorts_new:
                buckets[s] = []
            buckets["__others__"] = []
            unaffected = []  # 不属于该 parent_prefix 的条目（其他同级或无关），整体保持原有相对顺序
            for it in self.store.items:
                cat = it["category"] or "未分类"
                if cat.startswith(prefix_filter):
                    rest = cat[len(prefix_filter):]
                    if rest:
                        short = rest.split(SEPARATOR)[0].strip()
                        if short in buckets:
                            buckets[short].append(it)
                            continue
                unaffected.append(it)
            # 构造新前缀下的条目顺序
            affected = []
            for s in sibling_shorts_new:
                affected.extend(buckets[s])
            affected.extend(buckets["__others__"])
            # 回写：把 unaffected 中属于原范围的那些条目（即之前匹配 parent_prefix+SEPARATOR 的）替换为 affected；
            # 其他 unaffected 中原本不匹配的保留原有位置。
            # 简便做法：重新遍历原 items，凡命中 parent_prefix+SEPARATOR 的从 affected 依次取；否则保留。
            it_affected = iter(affected)
            new_items = []
            for it in self.store.items:
                cat = it["category"] or "未分类"
                if cat.startswith(prefix_filter):
                    rest = cat[len(prefix_filter):]
                    if rest:
                        short = rest.split(SEPARATOR)[0].strip()
                        if short in set(sibling_shorts_new):
                            try:
                                new_items.append(next(it_affected))
                                continue
                            except StopIteration:
                                pass
                new_items.append(it)
            # 若还有剩余 affected（理论上不会），追加
            for it in it_affected:
                new_items.append(it)
            self.store.items = new_items
        else:
            # 顶级：以 sibling_fulls_new (=顶级分类名) 作为顶级顺序
            top_names = list(sibling_fulls_new)
            buckets = OrderedDict((n, []) for n in top_names)
            others = []
            for it in self.store.items:
                cat = it["category"] or "未分类"
                root = cat.split(SEPARATOR)[0].strip()
                if root in buckets:
                    buckets[root].append(it)
                else:
                    others.append(it)
            new_items = []
            for n in top_names:
                new_items.extend(buckets[n])
            new_items.extend(others)
            self.store.items = new_items

        self.store.modified = True
        self._refresh_all()

        # 恢复窗口几何与 sash 位置，避免刷新后界面收缩
        self._restore_layout(geo, sash_pos)

    def on_cat_up(self):
        sel = self._selected_cat_fullpath()
        if not sel:
            return
        siblings, idx, parent, short = self._sibling_categories(sel)
        if idx <= 0:
            messagebox.showinfo("提示", f"'{short}' 已经是同级第一个。")
            return
        siblings[idx - 1], siblings[idx] = siblings[idx], siblings[idx - 1]
        self._reorder_sibling_categories(siblings, parent)
        self.set_status(f"分类 '{short}' 已上移。")

    def on_cat_down(self):
        sel = self._selected_cat_fullpath()
        if not sel:
            return
        siblings, idx, parent, short = self._sibling_categories(sel)
        if idx < 0 or idx >= len(siblings) - 1:
            messagebox.showinfo("提示", f"'{short}' 已经是同级最后一个。")
            return
        siblings[idx + 1], siblings[idx] = siblings[idx], siblings[idx + 1]
        self._reorder_sibling_categories(siblings, parent)
        self.set_status(f"分类 '{short}' 已下移。")

    def on_cat_top(self):
        sel = self._selected_cat_fullpath()
        if not sel:
            return
        siblings, idx, parent, short = self._sibling_categories(sel)
        siblings = [sel] + [s for s in siblings if s != sel]
        self._reorder_sibling_categories(siblings, parent)
        self.set_status(f"分类 '{short}' 已置顶。")

    def on_cat_bottom(self):
        sel = self._selected_cat_fullpath()
        if not sel:
            return
        siblings, idx, parent, short = self._sibling_categories(sel)
        siblings = [s for s in siblings if s != sel] + [sel]
        self._reorder_sibling_categories(siblings, parent)
        self.set_status(f"分类 '{short}' 已置底。")

    # 更多：顶级分类按名称 / 数量 排序；或按选中分类的同级批量排序
    def on_cat_sort(self, mode="name", direction=1):
        # 若有选中且非顶级，则按同级范围排序；否则按顶级排序
        sel = self._selected_cat_fullpath()
        parent = ""
        if sel and SEPARATOR in sel:
            # 对选中分类的同级进行排序
            siblings, idx, parent, short = self._sibling_categories(sel)
        else:
            # 顶级
            siblings = list(self._cat_root_items())
        # 构建辅助信息：{full: {count, name}}
        _, agg = self.store.category_counts()
        if parent:
            # 同 parent_prefix 下的 agg key 就是这些 siblings 的 full
            info = {s: {"name": s.rsplit(SEPARATOR, 1)[-1], "count": agg.get(s, 0)} for s in siblings}
        else:
            info = {s: {"name": s, "count": agg.get(s, 0)} for s in siblings}

        def key_fn(full):
            d = info.get(full, {})
            if mode == "name":
                return d.get("name", full)
            if mode == "count":
                return (d.get("count", 0), d.get("name", full))
            return full

        reverse = direction < 0
        siblings_sorted = sorted(siblings, key=key_fn, reverse=reverse)
        self._reorder_sibling_categories(siblings_sorted, parent)
        scope = f"顶级分类" if not parent else f"同级分类（{parent} 下）"
        mode_cn = {"name": "名称", "count": "数量"}.get(mode, mode)
        dir_cn = "倒序" if reverse else "正序"
        self.set_status(f"已按{mode_cn}{dir_cn}重新排序（{scope}）。")

    def on_links_in_cat_sort(self, field="name", direction=1):
        """把 items 按原分组稳定性保持，仅对同「网址分类目录」= 完全相同 的条目内部按 field 排序。"""
        import functools
        groups = OrderedDict()
        other = []
        for idx, it in enumerate(self.store.items):
            cat = it["category"] or "未分类"
            if cat not in groups:
                groups[cat] = []
            groups[cat].append(it)

        # 若用户选中了某个分类，则只对该分类（含其子孙）生效
        sel = self._selected_cat_fullpath()

        def need_sort(cat):
            if not sel:
                return True
            return cat == sel or cat.startswith(sel + SEPARATOR)

        reverse = direction < 0
        for cat, arr in groups.items():
            if need_sort(cat):
                if field == "name":
                    arr.sort(key=lambda x: (x["name"] or "").lower(), reverse=reverse)
                elif field == "url":
                    arr.sort(key=lambda x: (x["url"] or "").lower(), reverse=reverse)
                elif field == "desc":
                    arr.sort(key=lambda x: (x["desc"] or "").lower(), reverse=reverse)

        new_items = []
        for cat, arr in groups.items():
            new_items.extend(arr)

        # 保存窗口几何与 sash 位置
        geo, sash_pos = self._save_layout()

        self.store.items = new_items
        self.store.modified = True
        self._refresh_all()

        # 恢复窗口几何与 sash 位置
        self._restore_layout(geo, sash_pos)

        field_cn = {"name": "网站名称", "url": "链接", "desc": "描述"}.get(field, field)
        dir_cn = "倒序" if reverse else "正序"
        scope = f"选中分类 '{sel}' 下" if sel else "全局所有分类下"
        self.set_status(f"{scope}网址条目已按「{field_cn}」{dir_cn}排序。")

    def on_cat_original(self):
        """恢复为表格内按首次出现的原始顶级顺序（基本等同于导入时顺序）"""
        # 直接根据首次出现顺序重建顶级
        roots = self._cat_root_items()
        # _reorder_sibling_categories 用同级（顶级）重新分组即可
        self._reorder_sibling_categories(list(roots), "")
        self.set_status("已按导入时的原始表格顺序恢复分类排序。")

    # ---- 其它 ----
    def on_cat_select(self, _evt):
        self._refresh_table()

    def on_reset_filter(self):
        self.search_var.set("")
        root = self.cat_tree.get_children()
        if root:
            self.cat_tree.selection_set(root[0])
        self._refresh_table()
        self.set_status("已重置筛选。")

    # ---- 图标下载 / 更新 ----
    def _table_selected_store_indices(self):
        rowid_to_idx = {r: i for i, r in self._data_rowids}
        return [rowid_to_idx[r] for r in self.table.selection() if r in rowid_to_idx]

    def on_download_icons(self):
        if requests is None:
            messagebox.showerror("缺少依赖", "未安装 requests 库，无法下载图标。\n请运行：pip install requests")
            return
        sel = self._table_selected_store_indices()
        if sel:
            dlg = ScopeDialog(self, all_count=len(self.store.items), sel_count=len(sel))
            choice = dlg.result
            if choice is None:
                return
            idxs = list(range(len(self.store.items))) if choice == "all" else sel
        else:
            idxs = list(range(len(self.store.items)))
        # 「下载图标」跳过已存在的本地图标，「更新图标」强制重新下载覆盖
        self._download_icons(idxs, title="下载图标", skip_existing=True)

    def _download_icons(self, idxs, title="下载图标", skip_existing=False):
        # 去重（按 logo_key），保留顺序
        tasks = []
        seen = set()
        for i in idxs:
            it = self.store.items[i]
            key = logo_key(it.get("url", ""))
            if key and key not in seen:
                seen.add(key)
                tasks.append((i, key))
        if not tasks:
            messagebox.showinfo("提示", "没有有效的链接可下载。")
            return
        dlg = DownloadDialog(self, total=len(tasks), title=title)
        threading.Thread(
            target=self._download_worker, args=(tasks, dlg, skip_existing), daemon=True
        ).start()
        dlg.wait_window()
        skipped = getattr(dlg, "skip_count", 0)
        self.set_status(
            f"图标下载完成：成功 {dlg.ok_count} / 失败 {dlg.fail_count}"
            + (f" / 跳过 {skipped}（已存在）" if skipped else "")
            + f"（共 {len(tasks)}）。已保存到 {LOGO_DIR}"
        )

    def _download_worker(self, tasks, dlg, skip_existing=False):
        service = FaviconService(log_callback=lambda m: dlg.append_log(m))
        ok = fail = skipped = 0
        failed_keys = []
        for n, (i, key) in enumerate(tasks, start=1):
            if dlg.cancelled:
                dlg.append_log("用户已取消。")
                break
            if skip_existing and _logo_key_exists(key):
                skipped += 1
                service.log(f"跳过（已存在本地图标）: {key}")
                dlg.set_progress(n, len(tasks), key, None)
                self._apply_icon_to_cache(key)  # 已有本地图标，同步缓存路径
                continue
            success, _ = service.download_one(key)
            if success:
                ok += 1
                self._apply_icon_to_cache(key)  # 下载成功才改写缓存，失败不改写
            else:
                fail += 1
                failed_keys.append(key)
            dlg.set_progress(n, len(tasks), key, success)
            time.sleep(0.3)
        # 写出失败清单
        if failed_keys:
            try:
                with open(os.path.join(LOGO_DIR, "failed_downloads.txt"), "w", encoding="utf-8") as f:
                    f.write("域名/key\n")
                    for k in failed_keys:
                        f.write(k + "\n")
            except Exception:
                pass
        dlg.finish(ok, fail, skipped)

    def _apply_icon_to_cache(self, key):
        """下载成功（或本地已有）后，把缓存（store.items）中同域名条目的图标路径改写为本地图标。"""
        for it in self.store.items:
            if logo_key(it.get("url", "")) == key:
                rel = _logo_local_rel(it.get("url", ""))
                if rel:
                    it["icon"] = rel

    def _logo_dir_empty(self):
        """图标文件夹（assets/logo）是否没有任何图标文件。"""
        try:
            return not any(
                f.lower().endswith(ICON_EXTS) for f in os.listdir(LOGO_DIR)
            )
        except OSError:
            return True

    def on_update_icon(self):
        if self._logo_dir_empty():
            messagebox.showinfo(
                "提示",
                "图标文件夹（assets\\logo）为空，还没有可更新的图标。\n"
                "请直接使用「下载图标」下载，无需更新。",
            )
            return
        idxs = self._table_selected_store_indices()
        if not idxs:
            messagebox.showinfo("提示", "请先在右侧表格中选择要更新图标的行（可多选）。")
            return
        if len(idxs) == 1:
            self._icon_update_single(idxs[0])
        else:
            if messagebox.askyesno(
                "更新图标",
                f"将重新从网络下载选中的 {len(idxs)} 个站点图标（强制覆盖所有旧图标，已下载过的也会重新下载）？\n\n"
                "提示：选中单行时点此按钮可选择“本地图片”手动指定图标。",
            ):
                self._download_icons(idxs, title="更新图标", skip_existing=False)

    def _icon_update_single(self, idx):
        it = self.store.items[idx]
        name = it.get("name") or it.get("url") or "该站点"
        dlg = IconUpdateDialog(self, name)
        res = dlg.result
        if res == "web":
            self._download_icons([idx], title="更新图标")
        elif res == "local":
            self._pick_local_icon(idx)

    def _pick_local_icon(self, idx):
        path = filedialog.askopenfilename(
            title="选择图标图片",
            filetypes=[("图片", "*.png *.jpg *.jpeg *.gif *.webp *.svg *.ico"), ("所有文件", "*.*")],
        )
        if not path:
            return
        it = self.store.items[idx]
        key = logo_key(it.get("url", ""))
        if not key:
            messagebox.showerror("错误", "该条目没有有效链接，无法生成图标文件名。")
            return
        try:
            with open(path, "rb") as f:
                data = f.read()
        except Exception as e:
            messagebox.showerror("读取失败", str(e))
            return
        svc = FaviconService()
        try:
            svc.remove_existing(key)  # 清除旧图标，避免多扩展名并存
            name = svc.save_as_png(key, data)  # 统一转为 PNG
        except Exception as e:
            messagebox.showerror("保存失败", str(e))
            return
        rel = _logo_local_rel(it.get("url", ""))
        if rel:
            it["icon"] = rel  # 同步改写缓存中的图标路径
        self.set_status(f"已用本地图片更新图标：{os.path.basename(name)}（已统一为 PNG）")

    def set_status(self, msg):
        self.status_var.set(msg)


# ============== 新增 / 编辑 对话框 ==============
class ItemDialog(tk.Toplevel):
    def __init__(self, master, title, item=None):
        super().__init__(master)
        self.title(title)
        self.resizable(False, False)
        self.result = None
        self.transient(master)
        self.grab_set()

        frm = ttk.Frame(self, padding=12)
        frm.pack(fill=tk.BOTH, expand=True)

        self.var_name = tk.StringVar(value=item["name"] if item else "")
        self.var_cat = tk.StringVar(value=item["category"] if item else "")
        self.var_url = tk.StringVar(value=item["url"] if item else "")
        self.var_desc = tk.StringVar(value=item["desc"] if item else "")

        def row(r, label, widget):
            ttk.Label(frm, text=label).grid(row=r, column=0, sticky=tk.E, padx=4, pady=4)
            widget.grid(row=r, column=1, sticky=tk.W + tk.E, padx=4, pady=4)

        frm.columnconfigure(1, weight=1)
        row(0, "网站名称：", ttk.Entry(frm, textvariable=self.var_name, width=60))
        row(1, "网址分类目录：", ttk.Entry(frm, textvariable=self.var_cat, width=60))
        ttk.Label(frm, text="（多级用 '>' 分隔，如：实用工具>图片处理）",
                  foreground="#888").grid(row=1, column=2, padx=6, sticky=tk.W)
        row(2, "链接：", ttk.Entry(frm, textvariable=self.var_url, width=60))

        ttk.Label(frm, text="描述：").grid(row=3, column=0, sticky=tk.NE, padx=4, pady=4)
        desc_txt = tk.Text(frm, width=60, height=6)
        desc_txt.grid(row=3, column=1, sticky=tk.W + tk.E, padx=4, pady=4)
        if item and item["desc"]:
            desc_txt.insert("1.0", item["desc"])

        btns = ttk.Frame(frm)
        btns.grid(row=4, column=0, columnspan=3, pady=(12, 0), sticky=tk.E)

        def ok():
            name = self.var_name.get().strip()
            cat = self.var_cat.get().strip()
            url = self.var_url.get().strip()
            desc = desc_txt.get("1.0", tk.END).strip()
            if not name and not url:
                messagebox.showwarning("缺少内容", "网站名称与链接至少填写一项。", parent=self)
                return
            self.result = {"name": name, "category": cat, "url": url, "desc": desc}
            self.destroy()

        ttk.Button(btns, text="确定", command=ok).pack(side=tk.RIGHT, padx=4)
        ttk.Button(btns, text="取消", command=self.destroy).pack(side=tk.RIGHT)

        self.bind("<Return>", lambda e: ok())
        self.bind("<Escape>", lambda e: self.destroy())


# ============== 图标下载相关对话框 ==============
class ScopeDialog(tk.Toplevel):
    def __init__(self, master, all_count, sel_count):
        super().__init__(master)
        self.title("下载图标范围")
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()
        self.result = None
        frm = ttk.Frame(self, padding=14)
        frm.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frm, text="请选择要下载图标的范围：").pack(anchor=tk.W, pady=(0, 8))
        var = tk.StringVar(value="sel" if sel_count else "all")
        ttk.Radiobutton(frm, text=f"仅选中行（{sel_count} 个）", variable=var, value="sel").pack(anchor=tk.W, pady=2)
        ttk.Radiobutton(frm, text=f"全部网址（{all_count} 个）", variable=var, value="all").pack(anchor=tk.W, pady=2)
        btns = ttk.Frame(frm)
        btns.pack(anchor=tk.E, pady=(12, 0))
        ttk.Button(btns, text="开始", command=lambda: self._set(var.get())).pack(side=tk.RIGHT, padx=4)
        ttk.Button(btns, text="取消", command=self.destroy).pack(side=tk.RIGHT)

    def _set(self, v):
        self.result = v
        self.destroy()


class IconUpdateDialog(tk.Toplevel):
    def __init__(self, master, name):
        super().__init__(master)
        self.title("更新图标")
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()
        self.result = None
        frm = ttk.Frame(self, padding=14)
        frm.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frm, text=f"如何更新「{name}」的图标？").pack(anchor=tk.W, pady=(0, 8))
        ttk.Button(frm, text="重新下载网络图标", width=22, command=lambda: self._set("web")).pack(fill=tk.X, pady=3)
        ttk.Button(frm, text="选择本地图片文件", width=22, command=lambda: self._set("local")).pack(fill=tk.X, pady=3)
        ttk.Button(frm, text="取消", width=22, command=self.destroy).pack(fill=tk.X, pady=(8, 0))

    def _set(self, v):
        self.result = v
        self.destroy()


class DownloadDialog(tk.Toplevel):
    def __init__(self, master, total, title="下载图标"):
        super().__init__(master)
        self.title(title)
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()
        self.total = total
        self.cancelled = False
        self.ok_count = 0
        self.fail_count = 0
        self.skip_count = 0
        self.bind("<Destroy>", lambda e: setattr(self, "cancelled", True))
        frm = ttk.Frame(self, padding=12)
        frm.pack(fill=tk.BOTH, expand=True)
        self.label = ttk.Label(frm, text=f"准备下载 0 / {total}")
        self.label.pack(anchor=tk.W, pady=(0, 4))
        self.pb = ttk.Progressbar(frm, length=460, maximum=max(total, 1), mode="determinate")
        self.pb.pack(fill=tk.X, pady=(0, 6))
        self.log = tk.Text(frm, width=64, height=14, state="disabled")
        self.log.pack(fill=tk.BOTH, expand=True)
        btns = ttk.Frame(frm)
        btns.pack(anchor=tk.E, pady=(8, 0))
        self.btn_cancel = ttk.Button(btns, text="取消", command=self.on_cancel)
        self.btn_cancel.pack(side=tk.RIGHT, padx=4)
        self.btn_close = ttk.Button(btns, text="关闭", command=self.destroy, state="disabled")
        self.btn_close.pack(side=tk.RIGHT)

    def append_log(self, msg):
        if self.winfo_exists():
            self.after(0, self._append, msg)

    def _append(self, msg):
        if not self.winfo_exists():
            return
        self.log.configure(state="normal")
        self.log.insert(tk.END, msg + "\n")
        self.log.see(tk.END)
        self.log.configure(state="disabled")

    def set_progress(self, done, total, key, success):
        if self.winfo_exists():
            self.after(0, self._setp, done, total, key, success)

    def _setp(self, done, total, key, success):
        if not self.winfo_exists():
            return
        self.pb["value"] = done
        if success is None:
            tag = "→ 跳过"
        else:
            tag = "✓" if success else "✗"
        self.label.config(text=f"下载 {done} / {total}  [{tag} {key}]")

    def finish(self, ok, fail, skipped=0):
        if self.winfo_exists():
            self.after(0, self._finish, ok, fail, skipped)

    def _finish(self, ok, fail, skipped=0):
        if not self.winfo_exists():
            return
        self.ok_count, self.fail_count, self.skip_count = ok, fail, skipped
        skipped_txt = f" / 跳过 {skipped}" if skipped else ""
        self.label.config(text=f"完成：成功 {ok} / 失败 {fail}{skipped_txt}（共 {self.total}）")
        self.append_log(f"下载完成：成功 {ok} 个，失败 {fail} 个。"
                         + (f"跳过 {skipped} 个（已存在本地图标）。" if skipped else ""))
        if fail:
            self.append_log("失败项已记录到 assets/logo/failed_downloads.txt")
        self.btn_cancel.config(state="disabled")
        self.btn_close.config(state="normal")

    def on_cancel(self):
        self.cancelled = True
        self.append_log("已请求取消，等待当前任务结束…")


# ============== 启动 ==============
def main():
    app = App()

    # 如果命令行给了文件路径，自动导入（按扩展名区分 Excel / JSON）
    if len(sys.argv) >= 2 and os.path.isfile(sys.argv[1]):
        try:
            path = sys.argv[1]
            if path.lower().endswith(".json"):
                cnt = app.store.load_json(path)
            else:
                cnt = app.store.load_excel(path)
            app.set_status(f"已自动导入 {path} （共 {cnt} 条）")
            app._refresh_all()
            root = app.cat_tree.get_children()
            if root:
                app.cat_tree.selection_set(root[0])
                app._refresh_table()
        except Exception as e:
            app.set_status(f"自动导入失败: {e}")

    app.mainloop()


if __name__ == "__main__":
    main()
